import re
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Tuple, Any
from dataclasses import dataclass, field

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


@dataclass
class SubgroupInfo:
    name: str
    column: int


@dataclass
class GroupInfo:
    group_name: str
    subgroups: List[SubgroupInfo] = field(default_factory=list)


@dataclass
class Lesson:
    time: str
    subject: str
    teacher: str
    room: str
    type: str


@dataclass
class ScheduleDay:
    day_name: str
    lessons: List[Lesson] = field(default_factory=list)


# Longer prefixes first so "ТММ"/"ЭТ"/"КТ"/… win over shorter ones ("ТМ", "М").
ALLOWED_PREFIXES = [
    "ТММ",  # магистратура
    "ЭТ",
    "ЛТ",
    "КТ",
    "ТБ",
    "ТМ",
    "МР",
    "ИС",
    "НК",
    "И",
    "У",
    "П",
    "М",
]

# Spaced day titles in magistracy Excel: "П О Н Е Д Е Л Ь Н И К"
_MASTERS_DAY_MAP = {
    "понедельник": "Понедельник",
    "вторник": "Вторник",
    "среда": "Среда",
    "четверг": "Четверг",
    "пятница": "Пятница",
    "суббота": "Суббота",
    "воскресенье": "Воскресенье",
}

HOLIDAYS = {
    "01.01": "Новый год",
    "02.01": "Новогодние каникулы",
    "03.01": "Новогодние каникулы",
    "04.01": "Новогодние каникулы",
    "05.01": "Новогодние каникулы",
    "06.01": "Новогодние каникулы",
    "07.01": "Рождество Христово",
    "08.01": "Новогодние каникулы",
    "23.02": "День защитника Отечества",
    "08.03": "Международный женский день",
    "01.05": "Праздник Весны и Труда",
    "09.05": "День Победы",
    "12.06": "День России",
    "04.11": "День народного единства",
}

# 0-based row of first pair (числитель) for each weekday
DAY_START_ROWS = {
    "Понедельник": 4,
    "Вторник": 18,
    "Среда": 32,
    "Четверг": 46,
    "Пятница": 60,
    "Суббота": 74,
}

TIME_SLOTS = {
    1: "8:00-09:25",
    2: "09:35-11:00",
    3: "12:00-13:25",
    4: "13:35-15:00",
    5: "15:10-16:35",
    6: "17:45-19:10",
    7: "19:20-20:45",
}

LAST_ROW = 84


def get_semester_starts(now: Optional[datetime] = None) -> Tuple[datetime, datetime]:
    """Return (first_semester_start, second_semester_start) for the current academic year."""
    now = (now or datetime.now()).replace(hour=0, minute=0, second=0, microsecond=0)
    year = now.year
    if now.month >= 9:
        first = datetime(year, 9, 1)
        second = datetime(year + 1, 1, 13)
    else:
        first = datetime(year - 1, 9, 1)
        second = datetime(year, 1, 13)
    return first, second


def infer_course_from_group(group_name: str, now: Optional[datetime] = None) -> Optional[int]:
    """
    Derive course year from group code suffix (М-126 → 1, М-122 → 5 in 2026/27).
    """
    if not group_name:
        return None
    m = re.search(r"-(\d{2,3})\s*$", group_name.strip().upper())
    if not m:
        return None
    yy = int(m.group(1)[-2:])
    now = now or datetime.now()
    start_year = now.year if now.month >= 9 else now.year - 1
    entry = start_year % 100
    course = entry - yy + 1
    if 1 <= course <= 6:
        return course
    return None


def get_current_week_type(now: Optional[datetime] = None) -> str:
    now = (now or datetime.now()).replace(hour=0, minute=0, second=0, microsecond=0)
    first, second = get_semester_starts(now)
    semester_start = second if now >= second else first
    diff_days = max(0, (now - semester_start).days)
    weeks_diff = diff_days // 7
    return "Числитель" if weeks_diff % 2 == 0 else "Знаменатель"


def get_today_name() -> str:
    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
    return days[datetime.now().weekday()]


def get_date_for_day(day_name: str) -> datetime:
    day_map = {
        "Понедельник": 0,
        "Вторник": 1,
        "Среда": 2,
        "Четверг": 3,
        "Пятница": 4,
        "Суббота": 5,
        "Воскресенье": 6,
    }
    today = datetime.now()
    target_weekday = day_map.get(day_name, 0)
    current_weekday = today.weekday()
    diff = target_weekday - current_weekday
    return today + timedelta(days=diff)


def is_holiday(date: datetime) -> bool:
    return date.strftime("%d.%m") in HOLIDAYS


def get_holiday_name(date: datetime) -> str:
    return HOLIDAYS.get(date.strftime("%d.%m"), "Праздничный день")


def extract_group_name(text: Optional[str]) -> Optional[str]:
    if not text or not text.strip():
        return None
    cleaned = text.strip().upper().replace("\n", " ")
    # "НК-123 3 подгруппа", "М-126 (СПВ)", "И-125"
    for prefix in ALLOWED_PREFIXES:
        m = re.match(
            rf"^{re.escape(prefix)}\s*[-–]?\s*(\d{{2,3}})\b",
            cleaned,
        )
        if m:
            return f"{prefix}-{m.group(1)}"
    return None


def _roman_to_int(token: str) -> Optional[int]:
    """Parse simple Roman numerals I..X (also used in Excel: 'I подгруппа')."""
    t = token.lower().strip()
    table = {
        "i": 1,
        "ii": 2,
        "iii": 3,
        "iv": 4,
        "v": 5,
        "vi": 6,
        "vii": 7,
        "viii": 8,
        "ix": 9,
        "x": 10,
    }
    return table.get(t)


def extract_subgroup(text: Optional[str]) -> Optional[str]:
    if not text or not text.strip():
        return None
    lower = text.lower().replace("\n", " ").strip()
    # Arabic: "1 подгруппа", "2-я подгруппа"
    # Roman:  "I подгруппа", "II подгруппа" (common in real Excel from the uni)
    match = re.search(
        r"([ivxlcdm]+|\d+)\s*[-.]?\s*(?:я\s+)?подгрупп",
        lower,
    )
    if match:
        token = match.group(1)
        if token.isdigit():
            n = int(token)
        else:
            n = _roman_to_int(token)
        if n:
            return f"{n} подгруппа"
    match = re.search(r"подгрупп\w*\s*[:№]?\s*([ivxlcdm]+|\d+)", lower)
    if match:
        token = match.group(1)
        n = int(token) if token.isdigit() else _roman_to_int(token)
        if n:
            return f"{n} подгруппа"
    return None


def _column_looks_like_schedule_track(sheet: "SheetAdapter", col: int) -> bool:
    """True if this column has its own lesson text (not only inherited from a wide left merge)."""
    for row in range(4, min(LAST_ROW + 1, sheet.max_row or LAST_ROW + 1)):
        raw = sheet.get_cell_text(row, col)
        if raw and raw.strip() and raw.strip() not in ("-", "null"):
            return True
        # Own merge starting in this column (subgroup-specific block)
        merge = sheet.find_merge(row, col)
        if not merge:
            continue
        if merge["min_col"] != col:
            continue
        top = sheet.get_cell_text(merge["min_row"], merge["min_col"])
        if top and top.strip() and top.strip() not in ("-", "null"):
            return True
    return False


def _columns_have_different_lessons(
    sheet: "SheetAdapter", col_a: int, col_b: int
) -> bool:
    """
    True if col_b sometimes shows a different lesson than col_a
    (typical 1-я vs 2-я подгруппа of the same group).
    """
    for row in range(4, min(LAST_ROW + 1, sheet.max_row or LAST_ROW + 1)):
        raw_b = sheet.get_cell_text(row, col_b)
        if raw_b and raw_b.strip() and raw_b.strip() not in ("-", "null"):
            return True
        merge_b = sheet.find_merge(row, col_b)
        if merge_b and merge_b["min_col"] == col_b:
            top = sheet.get_cell_text(merge_b["min_row"], merge_b["min_col"])
            if top and top.strip() and top.strip() not in ("-", "null"):
                return True
        t_a = (sheet.get_text_with_merged(row, col_a) or "").strip()
        t_b = (sheet.get_text_with_merged(row, col_b) or "").strip()
        if t_b and t_a and t_b != t_a:
            # Shared stream merge would give the same text; difference ⇒ own track
            return True
        if t_b and not t_a:
            return True
    return False


def clean_subject(s: str) -> str:
    # Remove only standalone type markers, never mid-word (Электротехника).
    s = re.sub(
        r"(?i)\b(лекция|лекц\.?|лек\.?|практика|практ\.?|прак\.?|лабораторн\w*|лаб\.?|л/р)\b\.?",
        "",
        s,
    )
    # Keep «онлайн» in subject if it was on the title line; strip empty parens leftovers later.
    return re.sub(r"\s+", " ", s).strip(" .\t")


_ONLINE_RE = re.compile(r"онлайн|дистанц(?:ионн\w*)?|в\s*сдо|\bzoom\b|\bteams\b", re.I)


def _online_notes(text: str) -> List[str]:
    """Parenthetical notes that mention online / distance learning."""
    notes = []
    for m in re.finditer(r"\(([^)]+)\)", text):
        inner = m.group(1).strip()
        if _ONLINE_RE.search(inner):
            notes.append(inner)
    return notes


def _strip_online_markers(s: str) -> str:
    s = re.sub(r"\(\s*[^)]*онлайн[^)]*\)", " ", s, flags=re.I)
    s = re.sub(r"(?i)\bонлайн\b", " ", s)
    return re.sub(r"\s+", " ", s).strip(" ,.;")


def _is_pure_online_line(line: str) -> bool:
    low = line.lower().strip().strip("()")
    return bool(re.fullmatch(r"онлайн|дистанц(?:ионно|ионная форма)?", low))


def _room_token_from_line(line: str) -> Optional[str]:
    """
    Extract a real classroom token. Skip week-range lines like
    '(с 9 по 17 онлайн)' where the only digits are week numbers.
    """
    raw = line.strip()
    low = raw.lower()
    if _is_pure_online_line(raw):
        return "онлайн"
    # Pure week/online note without a classroom at the start
    if re.match(r"^\(?\s*с\s*\d+", low) and ("нед" in low or _ONLINE_RE.search(low)):
        return None
    if _ONLINE_RE.search(low) and "нед" in low and not re.match(r"^\d{2,4}", raw):
        return None

    # "234 (с 1 по 8 нед)" → room 234
    m = re.match(r"^(\d{2,4}[а-яА-Яa-zA-ZкКлЛ]?)\b", raw)
    if m:
        return m.group(1)

    # Otherwise take last room-like token, but not digits after «по» in week ranges
    cleaned = re.sub(r"с\s*\d+\s*по\s*\d+", " ", raw, flags=re.I)
    cleaned = re.sub(r"с\s*\d+\s*нед", " ", cleaned, flags=re.I)
    toks = re.findall(r"(\d{2,4}[а-яА-Яa-zA-ZкКлЛ]?)\b", cleaned)
    if not toks:
        return None
    return toks[-1]


def parse_lesson_text(text: str, time: str) -> Optional[Lesson]:
    lines = [line.strip() for line in text.replace("\r", "\n").split("\n") if line.strip()]
    if not lines:
        return None

    subject = clean_subject(lines[0])
    if not subject or subject in ("-", "null"):
        return None

    lesson_type = "занятие"
    text_lower = text.lower()
    if "лек" in text_lower:
        lesson_type = "лекция"
    elif "практ" in text_lower:
        lesson_type = "практика"
    elif "лаб" in text_lower or "л/р" in text_lower:
        lesson_type = "лабораторная"

    # Special whole-day / non-class activities
    upper = subject.upper()
    if "ВОЕНН" in upper or "ФИЗИЧЕСКАЯ КУЛЬТУРА" in upper or "ОБЩАЯ ФИЗИЧЕСКАЯ" in upper:
        if "лек" not in text_lower and "практ" not in text_lower and "лаб" not in text_lower:
            lesson_type = "занятие"

    online_notes = _online_notes(text)
    is_online = bool(online_notes) or bool(_ONLINE_RE.search(text))

    room = ""
    teacher = ""
    room_line_idx = None
    for i in range(len(lines) - 1, 0, -1):
        if _is_pure_online_line(lines[i]):
            room = "онлайн"
            room_line_idx = i
            teacher = ", ".join(_strip_online_markers(x) for x in lines[1:i] if x)
            teacher = re.sub(r"\s+,", ",", teacher).strip(" ,")
            break
        tok = _room_token_from_line(lines[i])
        if tok:
            room = tok
            room_line_idx = i
            # teacher = middle lines; strip online markers from teacher line
            teacher_parts = []
            for j in range(1, i):
                part = _strip_online_markers(lines[j])
                if part and not _is_pure_online_line(part):
                    teacher_parts.append(part)
            # same-line teacher before room
            before = _strip_online_markers(lines[i][: lines[i].find(tok)].strip(" ,"))
            if before and "нед" not in before.lower() and not _ONLINE_RE.search(before):
                teacher_parts.append(before)
            teacher = ", ".join(teacher_parts)
            break

    # Teacher on its own line with «(онлайн)» and no numeric room
    if not teacher:
        for j in range(1, len(lines)):
            if j == room_line_idx:
                continue
            if _is_pure_online_line(lines[j]):
                continue
            if "нед" in lines[j].lower() and _ONLINE_RE.search(lines[j]):
                continue
            cand = _strip_online_markers(lines[j])
            if cand and not re.fullmatch(r"\d{2,4}[а-яА-Яa-zA-ZкКлЛ]?", cand):
                teacher = cand
                break

    # Single-line: "ФИЗИЧЕСКАЯ КУЛЬТУРА и СПОРТ лек. 319"
    if not room and len(lines) == 1:
        m = re.search(r"(\d{2,4}[а-яА-Яa-zA-ZкКлЛ]?)\s*$", lines[0])
        if m and "нед" not in lines[0].lower():
            room = m.group(1)
            subject = clean_subject(lines[0][: m.start()])

    subject = _strip_online_markers(subject)

    # Make online visible in the room field (UI already shows «Аудитория: …»)
    if is_online:
        note = "; ".join(online_notes) if online_notes else "онлайн"
        # shorten pure online
        if re.fullmatch(r"онлайн", note, re.I):
            note = "онлайн"
        if not room or room.lower() == "онлайн":
            room = note if note else "онлайн"
        elif "онлайн" not in room.lower():
            room = f"{room} · {note}"

    return Lesson(time=time, subject=subject, teacher=teacher, room=room, type=lesson_type)


class SheetAdapter:
    """
    Unified adapter for openpyxl and xlrd.

    All coordinates are 0-based inclusive.
    Merged ranges are stored as 0-based inclusive min/max.
    """

    def __init__(self, sheet: Any, workbook: Any, is_xls: bool = False):
        self.sheet = sheet
        self.workbook = workbook
        self.is_xls = is_xls
        # list of dicts: min_row, max_row, min_col, max_col (0-based inclusive)
        self.merged_ranges: List[Dict[str, int]] = []

        if is_xls:
            # xlrd: (rlo, rhi, clo, chi) half-open, 0-based
            for rng in sheet.merged_cells:
                rlo, rhi, clo, chi = rng
                if rhi <= rlo or chi <= clo:
                    continue
                self.merged_ranges.append({
                    "min_row": rlo,
                    "max_row": rhi - 1,
                    "min_col": clo,
                    "max_col": chi - 1,
                })
        else:
            # openpyxl: 1-based inclusive
            for rng in sheet.merged_cells.ranges:
                self.merged_ranges.append({
                    "min_row": rng.min_row - 1,
                    "max_row": rng.max_row - 1,
                    "min_col": rng.min_col - 1,
                    "max_col": rng.max_col - 1,
                })

    def get_cell_text(self, row: int, col: int) -> Optional[str]:
        try:
            if self.is_xls:
                if row < 0 or col < 0 or row >= self.sheet.nrows or col >= self.sheet.ncols:
                    return None
                val = self.sheet.cell(row, col).value
            else:
                cell = self.sheet.cell(row=row + 1, column=col + 1)
                val = cell.value

            if val is None:
                return None
            s = str(val).strip()
            return s if s else None
        except (IndexError, TypeError, AttributeError):
            return None

    def find_merge(self, row: int, col: int) -> Optional[Dict[str, int]]:
        for merge in self.merged_ranges:
            if (
                merge["min_row"] <= row <= merge["max_row"]
                and merge["min_col"] <= col <= merge["max_col"]
            ):
                return merge
        return None

    def get_text_with_merged(self, row: int, col: int) -> Optional[str]:
        """Return cell text, resolving merged cells (value lives in top-left)."""
        raw = self.get_cell_text(row, col)
        if raw and raw not in ("-", "null"):
            return raw

        merge = self.find_merge(row, col)
        if merge:
            top_left = self.get_cell_text(merge["min_row"], merge["min_col"])
            if top_left and top_left not in ("-", "null"):
                return top_left
        return None

    @property
    def max_column(self) -> int:
        if self.is_xls:
            return self.sheet.ncols
        return self.sheet.max_column or 0

    @property
    def max_row(self) -> int:
        if self.is_xls:
            return self.sheet.nrows
        return self.sheet.max_row or 0


def open_workbook(file_path: str) -> Tuple[SheetAdapter, Any]:
    """Open Excel file. Files from dksta are often real .xls saved as .xlsx."""
    errors = []

    # 1) True .xls (Composite Document) — need xlrd + formatting_info for merges
    if HAS_XLRD:
        try:
            wb = xlrd.open_workbook(file_path, formatting_info=True)
            sheet = wb.sheet_by_index(0)
            return SheetAdapter(sheet, wb, is_xls=True), wb
        except Exception as e:
            errors.append(f"xlrd+fmt: {e}")
            try:
                wb = xlrd.open_workbook(file_path)
                sheet = wb.sheet_by_index(0)
                return SheetAdapter(sheet, wb, is_xls=True), wb
            except Exception as e2:
                errors.append(f"xlrd: {e2}")

    # 2) Real .xlsx
    if HAS_OPENPYXL:
        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active
            return SheetAdapter(sheet, wb, is_xls=False), wb
        except Exception as e:
            errors.append(f"openpyxl: {e}")

    raise ValueError(
        f"Cannot open {file_path}. Install openpyxl/xlrd. Details: {'; '.join(errors)}"
    )


def close_workbook(adapter: SheetAdapter, wb: Any) -> None:
    if not adapter.is_xls:
        try:
            wb.close()
        except Exception:
            pass


def _group_header_span(
    sheet: SheetAdapter, col: int, group_row: int = 2
) -> Tuple[int, int]:
    """Return (start_col, end_col) for group name cell / merge on group_row."""
    start_col = col
    end_col = col
    merge = sheet.find_merge(group_row, col)
    if merge and merge["min_row"] <= group_row <= merge["max_row"]:
        start_col = merge["min_col"]
        end_col = merge["max_col"]
    return start_col, end_col


def _detect_group_header_row(sheet: SheetAdapter) -> int:
    """
    АиЭ: group names on row 2; some МТФ on row 1; магистратура (ТММ-…) often on row 4.
    Pick the row (1..6) with more recognized group names.
    """
    best_row, best_count = 2, 0
    for row in range(1, 7):
        seen = set()
        count = 0
        for col in range(sheet.max_column):
            text = sheet.get_cell_text(row, col)
            name = extract_group_name(text)
            if name is None:
                name = extract_group_name(sheet.get_text_with_merged(row, col))
            if name and name not in seen:
                seen.add(name)
                count += 1
        if count > best_count:
            best_row, best_count = row, count
    return best_row


def _normalize_spaced_day(raw: str) -> Optional[str]:
    """'П О Н Е Д Е Л Ь Н И К' / 'ПОНЕДЕЛЬНИК' → 'Понедельник'."""
    if not raw:
        return None
    compact = re.sub(r"\s+", "", raw).lower().replace("ё", "е")
    return _MASTERS_DAY_MAP.get(compact)


def _normalize_masters_time(raw: str) -> str:
    """'17.45  -  19.10' / '19.20\\n-\\n20-45' → '17:45-19:10'."""
    if not raw:
        return ""
    s = raw.replace("\n", " ").replace("\r", " ")
    s = s.replace(".", ":").replace("—", "-").replace("–", "-")
    s = re.sub(r"\s+", "", s)
    # '20-45' → '20:45' when used as mm part
    s = re.sub(r"(\d{1,2})-(\d{2})(?=\d|\-|:|$)", r"\1:\2", s)
    # After first replace we may get 17:45-19:10 or 17:45-:19:10 artifacts — clean
    s = s.replace("::", ":")
    m = re.search(r"(\d{1,2}:\d{2}).*?(\d{1,2}:\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return re.sub(r"\s+", " ", raw).strip()


def looks_like_masters_sheet(sheet: SheetAdapter) -> bool:
    """
    True for magistracy evening grid (ТММ-…).

    Important: bachelor Excel (АиЭ/МТФ) also uses spaced weekday titles
    like «П О Н Е Д Е Л Ь Н И К» — that alone must NOT trigger masters mode,
    otherwise subgroup detection breaks (extra «3 подгруппа» etc.).
    """
    has_tmm = False
    has_classic_header_group = False
    for row in range(min(8, sheet.max_row or 0)):
        for col in range(min(sheet.max_column or 0, 40)):
            t = sheet.get_cell_text(row, col) or ""
            name = extract_group_name(t) or extract_group_name(
                sheet.get_text_with_merged(row, col)
            )
            if not name:
                continue
            if name.startswith("ТММ"):
                has_tmm = True
            elif row <= 3:
                # И-/У-/М-/КТ-… in the usual header rows ⇒ bachelor sheet
                has_classic_header_group = True
    if has_tmm:
        return True
    # Fallback: compact sheet without classic groups (rare)
    if not has_classic_header_group and (sheet.max_row or 0) <= 45:
        for row in range(min(12, sheet.max_row or 0)):
            if _normalize_spaced_day(sheet.get_cell_text(row, 0) or ""):
                return True
    return False


def parse_masters_groups(sheet: SheetAdapter) -> Dict[str, GroupInfo]:
    """
    Magistracy header:
      row with ТММ-125
      next row: I подгруппа | II подгруппа under lesson columns
    """
    groups: Dict[str, GroupInfo] = {}
    group_row = _detect_group_header_row(sheet)
    sub_row = group_row + 1

    for col in range(sheet.max_column):
        name = extract_group_name(sheet.get_cell_text(group_row, col))
        if name is None:
            name = extract_group_name(sheet.get_text_with_merged(group_row, col))
        if not name:
            continue
        if name in groups:
            continue
        info = GroupInfo(group_name=name)
        # Collect unique subgroup tracks on the next row (I / II …)
        for c in range(sheet.max_column):
            sub = extract_subgroup(sheet.get_text_with_merged(sub_row, c)) or extract_subgroup(
                sheet.get_cell_text(sub_row, c)
            )
            if not sub:
                continue
            if any(s.name == sub for s in info.subgroups):
                continue
            info.subgroups.append(SubgroupInfo(name=sub, column=c))
        if not info.subgroups:
            info.subgroups.append(SubgroupInfo(name="1 подгруппа", column=col))
        groups[name] = info
    return groups


def parse_masters_schedule_for_group(
    sheet: SheetAdapter,
    group_name: str,
    subgroup: Optional[str],
    week_type: str,
) -> List[ScheduleDay]:
    """
    Evening magistracy grid: day title in col0, pair#, time, week label, then subgroup cols.
    """
    groups = parse_masters_groups(sheet)
    info = groups.get(group_name)
    if not info or not info.subgroups:
        return []

    # Resolve subgroup column
    col = info.subgroups[0].column
    if subgroup:
        for s in info.subgroups:
            if s.name == subgroup or subgroup in s.name or s.name in subgroup:
                col = s.column
                break
        # also accept "1 подгруппа" vs "I подгруппа"
        want = subgroup.lower().replace("ё", "е")
        for s in info.subgroups:
            if "1" in want and ("1" in s.name or s.name.lower().startswith("i")):
                col = s.column
                break
            if "2" in want and ("2" in s.name or "ii" in s.name.lower()):
                col = s.column
                break

    week_key = "числитель" if week_type == "Числитель" else "знаменатель"
    by_day: Dict[str, List[Lesson]] = {d: [] for d in _MASTERS_DAY_MAP.values() if d != "Воскресенье"}
    current_day: Optional[str] = None

    for row in range(sheet.max_row or 0):
        day_raw = sheet.get_cell_text(row, 0) or ""
        day = _normalize_spaced_day(day_raw)
        if day:
            current_day = day

        week_raw = (sheet.get_cell_text(row, 3) or sheet.get_text_with_merged(row, 3) or "").strip().lower()
        if week_key not in week_raw:
            continue
        if not current_day:
            continue

        text = sheet.get_text_with_merged(row, col) or sheet.get_cell_text(row, col) or ""
        if not text.strip() or text.strip() in ("-", "null"):
            continue

        time_raw = sheet.get_cell_text(row, 2) or sheet.get_text_with_merged(row, 2) or ""
        # Sometimes time only on числитель row — look one row up
        if not time_raw.strip() and row > 0:
            time_raw = sheet.get_cell_text(row - 1, 2) or ""
        time = _normalize_masters_time(time_raw)
        lesson = parse_lesson_text(text, time or time_raw.strip())
        if lesson:
            by_day.setdefault(current_day, []).append(lesson)

    order = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    return [ScheduleDay(day_name=d, lessons=by_day.get(d, [])) for d in order]


def parse_groups_from_header(sheet: SheetAdapter) -> Dict[str, GroupInfo]:
    """
    Header:
      group_row — group names (may be merged across subgroup columns)
      group_row+1 — optional "N подгруппа"
    Same group can appear in several blocks (e.g. main + 3-я подгруппа).

    Important: for И-125 the name is often merged over cols of 1-я and 2-я
    subgroup; both labels on row 3 must become separate tracks.
    """
    groups: Dict[str, GroupInfo] = {}
    last_col = sheet.max_column
    group_row = _detect_group_header_row(sheet)
    sub_row = group_row + 1

    # Collect ordered group anchors: (start_col, end_col, group_name)
    anchors: List[Tuple[int, int, str]] = []
    seen_cols = set()

    for col in range(last_col):
        if col in seen_cols:
            continue

        text = sheet.get_cell_text(group_row, col)
        group_name = extract_group_name(text)
        if group_name is None:
            merged_text = sheet.get_text_with_merged(group_row, col)
            group_name = extract_group_name(merged_text)

        if group_name is None:
            continue

        start_col, end_col = _group_header_span(sheet, col, group_row=group_row)
        for c in range(start_col, end_col + 1):
            seen_cols.add(c)
        anchors.append((start_col, end_col, group_name))

    # МТФ 4–5: "5 курс" on group_row, actual name (М-122) on the next row
    for col in range(last_col):
        if col in seen_cols:
            continue
        label = (sheet.get_cell_text(group_row, col) or "").strip().lower()
        if "курс" not in label and not extract_group_name(
            sheet.get_cell_text(group_row, col)
        ):
            # still try subgroup row for orphan group names
            pass
        name = extract_group_name(sheet.get_cell_text(sub_row, col))
        if name is None:
            name = extract_group_name(sheet.get_text_with_merged(sub_row, col))
        if name is None:
            continue
        if any(a[2] == name for a in anchors):
            continue
        start_col, end_col = col, col
        merge = sheet.find_merge(sub_row, col)
        if merge and merge["min_row"] <= sub_row <= merge["max_row"]:
            start_col, end_col = merge["min_col"], merge["max_col"]
        for c in range(start_col, end_col + 1):
            seen_cols.add(c)
        anchors.append((start_col, end_col, name))

    anchors.sort(key=lambda a: a[0])

    for idx, (start_col, end_col, group_name) in enumerate(anchors):
        if idx + 1 < len(anchors):
            block_end = anchors[idx + 1][0] - 1
        else:
            block_end = last_col - 1
            for c in range(end_col + 1, last_col):
                t = (
                    sheet.get_cell_text(group_row, c)
                    or sheet.get_text_with_merged(group_row, c)
                    or ""
                )
                tl = t.strip().lower().replace("\n", " ")
                if tl in ("время", "№ пары", "день недели", "день недели"):
                    block_end = c - 1
                    break

        block_end = max(block_end, end_col)

        if group_name not in groups:
            groups[group_name] = GroupInfo(group_name=group_name)

        found: List[Tuple[str, int]] = []

        def add_track(name: str, c: int) -> None:
            if any(s[0] == name for s in found):
                return
            if any(s[1] == c for s in found):
                return
            found.append((name, c))

        # 1) Explicit labels on subgroup row (resolve merges — "2 подгруппа" may sit in a merge)
        for c in range(start_col, block_end + 1):
            sub_text = sheet.get_text_with_merged(sub_row, c)
            if not sub_text:
                sub_text = sheet.get_cell_text(sub_row, c)
            # also plain group_row cell if subgroup written into group header
            if not sub_text:
                raw2 = sheet.get_cell_text(group_row, c)
                if raw2 and extract_subgroup(raw2):
                    sub_text = raw2
            subgroup = extract_subgroup(sub_text)
            if subgroup:
                add_track(subgroup, c)

        # 2) Multi-column group header (e.g. И-125 over cols 13–15) but only "1 подгруппа"
        #    found — look for a second schedule track under the same header.
        if end_col > start_col and len(found) < 2:
            for c in range(start_col, end_col + 1):
                if any(col == c for _, col in found):
                    continue
                if c != start_col and _column_looks_like_schedule_track(sheet, c):
                    used_nums = set()
                    for name, _ in found:
                        m = re.search(r"(\d)", name)
                        if m:
                            used_nums.add(int(m.group(1)))
                    n = 1
                    while n in used_nums:
                        n += 1
                    add_track(f"{n} подгруппа", c)

        # 2b) Classic layout: group name merged across 3 columns (1-я | middle | 2-я),
        #     but label "2 подгруппа" missing/unreadable — still expose rightmost col.
        if end_col - start_col >= 2 and len(found) == 1:
            second_col = end_col
            if not any(col == second_col for _, col in found):
                add_track("2 подгруппа", second_col)

        # 2c) Multi-col group header (И-125 over 13–15), only one track so far
        if len(found) == 1 and end_col > start_col:
            base_col = found[0][1]
            for c in (end_col, start_col + 2, end_col - 1):
                if c <= start_col or c > end_col:
                    continue
                if any(col == c for _, col in found):
                    continue
                if _column_looks_like_schedule_track(sheet, c) or _columns_have_different_lessons(
                    sheet, base_col, c
                ):
                    add_track("2 подгруппа", c)
                    break
            if len(found) == 1 and end_col != base_col:
                add_track("2 подгруппа", end_col)

        # 2d) Catch "2 подгруппа" labels that sit under empty group-name cells
        #     (between this group's start and the next group anchor).
        if len(found) < 2:
            for c in range(start_col + 1, block_end + 1):
                if any(col == c for _, col in found):
                    continue
                gn = extract_group_name(sheet.get_text_with_merged(group_row, c))
                if gn and gn != group_name:
                    break
                sub = extract_subgroup(sheet.get_text_with_merged(sub_row, c))
                if sub:
                    add_track(sub, c)

        # 2e) Informatics groups (И-xxx) almost always have 1-я and 2-я columns
        #     (name merge over 3 cols, or 2-я label two columns to the right).
        if len(found) == 1 and group_name.startswith("И"):
            for c in (start_col + 2, start_col + 1, end_col, start_col + 3):
                if c <= start_col or c > block_end:
                    continue
                if any(col == c for _, col in found):
                    continue
                add_track("2 подгруппа", c)
                break

        # 3) Still nothing — single column for the group
        if not found:
            add_track("1 подгруппа", start_col)

        for name, c in found:
            # Skip if this column already registered for the group
            if any(s.column == c for s in groups[group_name].subgroups):
                continue
            if not any(s.name == name for s in groups[group_name].subgroups):
                groups[group_name].subgroups.append(SubgroupInfo(name=name, column=c))
            else:
                # Name already used (another block) — keep column under a free label
                used = {s.name for s in groups[group_name].subgroups}
                for n in range(1, 8):
                    candidate = f"{n} подгруппа"
                    if candidate not in used:
                        groups[group_name].subgroups.append(
                            SubgroupInfo(name=candidate, column=c)
                        )
                        break

    for info in groups.values():
        by_col: Dict[int, SubgroupInfo] = {}
        for s in info.subgroups:
            if s.column not in by_col:
                by_col[s.column] = s
        info.subgroups = sorted(by_col.values(), key=lambda s: (s.name, s.column))

    return groups


def _pair_index_for_row(start_row: int, row: int) -> Optional[int]:
    """Return 0-based pair index if row belongs to this day block."""
    if row < start_row:
        return None
    offset = row - start_row
    if offset < 0 or offset >= 14:
        return None
    return offset // 2


def _time_range_for_pairs(first_pair: int, last_pair: int) -> str:
    """first_pair/last_pair are 0-based pair indices within the day."""
    start = TIME_SLOTS.get(first_pair + 1, "Неизвестно")
    end = TIME_SLOTS.get(last_pair + 1, "Неизвестно")
    if first_pair == last_pair:
        return start
    start_t = start.split("-")[0].strip()
    end_t = end.split("-")[-1].strip()
    return f"{start_t}-{end_t}"


def parse_lessons_for_day(
    sheet: SheetAdapter,
    start_row: int,
    column: int,
    week_type: str,
) -> List[Lesson]:
    """
    For each pair slot there are two rows:
      num_row = start_row + i*2      (числитель)
      den_row = num_row + 1          (знаменатель)

    Cases handled via merges:
    - num+den merged → same lesson both weeks
    - only num or only den filled → that week only
    - horizontal merge across subgroups / whole stream → all those columns share text
    - vertical merge across many pairs (e.g. военная подготовка) → one lesson for the range
    """
    lessons: List[Lesson] = []
    # skip pair indices already covered by a multi-pair merge
    covered_pairs: set = set()

    for i in range(7):
        if i in covered_pairs:
            continue

        num_row = start_row + i * 2
        den_row = num_row + 1
        if num_row > LAST_ROW:
            break

        selected_row = num_row if week_type == "Числитель" else den_row
        if selected_row > LAST_ROW:
            break

        text = sheet.get_text_with_merged(selected_row, column)
        if not text or not text.strip() or text.strip() in ("-", "null"):
            # Fallback: if this week cell is empty but num+den are vertically
            # merged, get_text_with_merged already handled it. If only the other
            # week has a lesson, correctly return nothing.
            continue

        merge = sheet.find_merge(selected_row, column)
        first_pair = i
        last_pair = i

        if merge:
            # Pairs of this day touched by the merge
            for j in range(7):
                pair_num = start_row + j * 2
                pair_den = pair_num + 1
                # overlap of [pair_num, pair_den] with [min_row, max_row]
                if pair_den >= merge["min_row"] and pair_num <= merge["max_row"]:
                    first_pair = min(first_pair, j)
                    last_pair = max(last_pair, j)

            # Only collapse when merge covers more than one pair slot
            if last_pair > first_pair:
                for j in range(first_pair, last_pair + 1):
                    covered_pairs.add(j)
                # emit once at the first pair of the block
                if i != first_pair:
                    # re-process from first_pair next... but we're past it; emit now
                    pass

        time = _time_range_for_pairs(first_pair, last_pair)
        lesson = parse_lesson_text(text, time)
        if lesson:
            lessons.append(lesson)

        if last_pair > first_pair:
            for j in range(first_pair, last_pair + 1):
                covered_pairs.add(j)

    return lessons


def schedule_to_dict(schedule: List[ScheduleDay]) -> list:
    return [
        {
            "dayName": day.day_name,
            "lessons": [
                {
                    "time": l.time,
                    "subject": l.subject,
                    "teacher": l.teacher,
                    "room": l.room,
                    "type": l.type,
                }
                for l in day.lessons
            ],
        }
        for day in schedule
    ]


def parse_schedule_for_group(
    file_path: str,
    group_name: str,
    subgroup: Optional[str],
    week_type: Optional[str] = None,
) -> List[ScheduleDay]:
    sheet, wb = open_workbook(file_path)
    try:
        active_week = week_type or get_current_week_type()
        if looks_like_masters_sheet(sheet):
            return parse_masters_schedule_for_group(
                sheet, group_name, subgroup, active_week
            )

        groups = parse_groups_from_header(sheet)
        # normalize group name lookup
        lookup = group_name.strip().upper().replace(" ", "")
        info = groups.get(lookup) or groups.get(group_name)
        if info is None:
            # fuzzy: strip spaces
            for name, g in groups.items():
                if name.replace(" ", "") == lookup:
                    info = g
                    break
        if info is None:
            return []

        subgroup_column = None
        if subgroup:
            for s in info.subgroups:
                if s.name == subgroup or s.name.replace(" ", "") == subgroup.replace(" ", ""):
                    subgroup_column = s.column
                    break
        if subgroup_column is None and info.subgroups:
            subgroup_column = info.subgroups[0].column

        if subgroup_column is None:
            return []

        result: List[ScheduleDay] = []

        for day_name, start_row in DAY_START_ROWS.items():
            if start_row > LAST_ROW:
                break
            date = get_date_for_day(day_name)
            if is_holiday(date):
                h = get_holiday_name(date)
                result.append(
                    ScheduleDay(
                        day_name=day_name,
                        lessons=[
                            Lesson(time="", subject=h, teacher="", room="", type="праздник")
                        ],
                    )
                )
                continue
            lessons = parse_lessons_for_day(sheet, start_row, subgroup_column, active_week)
            result.append(ScheduleDay(day_name=day_name, lessons=lessons))

        return result
    finally:
        close_workbook(sheet, wb)


def get_groups_from_file(file_path: str) -> Dict[str, List[str]]:
    sheet, wb = open_workbook(file_path)
    try:
        if looks_like_masters_sheet(sheet):
            groups = parse_masters_groups(sheet)
        else:
            groups = parse_groups_from_header(sheet)
        return {name: [s.name for s in info.subgroups] for name, info in groups.items()}
    finally:
        close_workbook(sheet, wb)
